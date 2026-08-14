import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

SRC = Path(r'D:\CC\DB\data\Fund_IRG_updated_v2.xlsx')
OUT = Path(r'D:\CC\Mid\估值')

# 读取数据
wb = load_workbook(SRC, data_only=True)
ws = wb['申万二级行业分布']

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    manager = r[0]
    pct = r[3]
    if manager == '实盘主动' and pct is not None and pct > 0:
        rows.append(r)

rows.sort(key=lambda r: r[3] if r[3] else 0, reverse=True)

# 准备数据
labels = []
pct_vals = []
pb_vals = []
pb_pct_vals = []
pe_vals = []
pe_pct_vals = []
method_vals = []

for r in rows:
    ind2_name = f"{r[1]}-{r[2]}" if r[1] else str(r[2])  # 一级-二级
    labels.append(ind2_name)
    pct_vals.append(r[3] if r[3] else 0)
    pb_vals.append(r[4] if r[4] else None)
    pb_pct_vals.append(r[5] if r[5] else None)
    pe_vals.append(r[6] if r[6] else None)
    pe_pct_vals.append(r[7] if r[7] else None)
    method_vals.append(r[8])  # PB or PE

n = len(labels)
x = np.arange(n)

fig, axes = plt.subplots(2, 1, figsize=(28, 26))
colors_pb = ['steelblue' if m == 'PB' else 'lightgray' for m in method_vals]
colors_pe = ['darkorange' if m == 'PE' else 'lightgray' for m in method_vals]

# 子图1: PB历史百分位
ax1 = axes[0]
bars1 = ax1.bar(x, [v if v else 0 for v in pb_pct_vals], color=colors_pb, edgecolor='white')
ax1.set_title('PB历史百分位(%) (按市值占比降序)', fontsize=16, fontweight='bold')
ax1.set_xticks(x)
ax1.set_xticklabels(labels, rotation=90, fontsize=7)
ax1.set_ylabel('PB百分位(%)')
ax1.axhline(y=50, color='red', linestyle='--', alpha=0.5, label='50%分位线')
ax1.axhline(y=80, color='orange', linestyle='--', alpha=0.5, label='80%分位线')
ax1.legend(fontsize=9)
ax1.grid(axis='y', alpha=0.3)

# 子图2: PE历史百分位
ax2 = axes[1]
bars2 = ax2.bar(x, [v if v else 0 for v in pe_pct_vals], color=colors_pe, edgecolor='white')
ax2.set_title('PE历史百分位(%) (按市值占比降序)', fontsize=16, fontweight='bold')
ax2.set_xticks(x)
ax2.set_xticklabels(labels, rotation=90, fontsize=7)
ax2.set_ylabel('PE百分位(%)')
ax2.axhline(y=50, color='red', linestyle='--', alpha=0.5, label='50%分位线')
ax2.axhline(y=80, color='orange', linestyle='--', alpha=0.5, label='80%分位线')
ax2.legend(fontsize=9)
ax2.grid(axis='y', alpha=0.3)

# 添加图例说明
fig.text(0.5, 0.01, '蓝色=PB估值行业  橙色=PE估值行业  虚线=50%/80%分位参考线',
         ha='center', fontsize=11, color='gray')

fig.suptitle('实盘主动 申万二级行业估值百分位分布\n(按市值占比从大到小排序)', fontsize=18, fontweight='bold', y=0.99)
fig.tight_layout(rect=[0, 0.03, 1, 0.97])

out_path = OUT / 'sw2_valuation_chart.png'
fig.savefig(out_path, dpi=150, bbox_inches='tight')
plt.close()
print(f'[OK] {out_path}')
print(f'共 {n} 个行业')
