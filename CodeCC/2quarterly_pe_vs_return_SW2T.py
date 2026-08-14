r"""
季度估值 vs 次季度收益率 — 一级行业一个 Excel，每个二级行业一个 sheet
数据源：
  - D:\CC\DB\data\datayes_all_SW_Industries_Level2_mapped.csv  (估值 PE)
  - D:\CC\DB\data\priceSW2T.xlsx                                 (价格序列)
输出：
  - output/pe_vs_return/  每个一级行业一个 .xlsx 文件，每个二级行业一个 sheet
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpxlImage
from openpyxl.utils import get_column_letter
import os
import io
import tempfile

# ============ 路径 ============
VAL_PATH  = r"D:\CC\DB\data\datayes_all_SW_Industries_Level2_mapped.csv"
PRICE_PATH = r"D:\CC\DB\data\priceSW2T.xlsx"
OUT_DIR    = r"D:\CC\Mid\估值\REStemp"
OUT_CSV    = r"D:\CC\Mid\估值\REStemp\quarterly_pe_vs_return_data.csv"

os.makedirs(OUT_DIR, exist_ok=True)

# 中文字体
plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

# ============ 1. 读取估值数据 ============
print("读取估值数据...")
df_val = pd.read_csv(VAL_PATH, parse_dates=["tradeDate"])
df_val = df_val[["tradeDate", "pe", "Fetched_Ind_Name", "申万一级行业"]].copy()
df_val = df_val.dropna(subset=["pe"])
df_val["pe"] = pd.to_numeric(df_val["pe"], errors="coerce")
df_val = df_val[df_val["pe"] > 0]  # 过滤无效 PE
# 统一为 date-only，便于后续比较
df_val["tradeDate"] = df_val["tradeDate"].dt.date
df_val["tradeDate"] = pd.to_datetime(df_val["tradeDate"])

# ============ 2. 读取价格数据 ============
print("读取价格数据...")
df_price = pd.read_excel(PRICE_PATH, header=None)

# Row 1 = 行业名称, Row 2+ = 数据
ind_names = []
ind_codes  = []
for i in range(1, df_price.shape[1]):
    val = df_price.iloc[1, i]
    if pd.notna(val):
        parts = str(val).split("\n")
        ind_names.append(parts[0])
        ind_codes.append(parts[1] if len(parts) > 1 else "")
    else:
        ind_names.append("")
        ind_codes.append("")

# 构建价格 DataFrame
dates = pd.to_datetime(df_price.iloc[2:, 0], errors="coerce")
price_data = {}
for i in range(df_price.shape[1] - 1):
    col_data = pd.to_numeric(df_price.iloc[2:, i + 1], errors="coerce")
    if ind_names[i]:
        price_data[ind_names[i]] = col_data.values

df_prices = pd.DataFrame(price_data, index=dates.values)
df_prices.index = pd.to_datetime(df_prices.index)
# 统一为 date-only
df_prices.index = pd.to_datetime(df_prices.index.date)
df_prices = df_prices.sort_index()

# ============ 3. 确定季度末日期 ============
print("确定季度区间...")
all_val_dates = sorted(df_val["tradeDate"].unique())
all_price_dates = sorted(df_prices.index)

def get_quarter_end_dates(dates):
    """从日期列表中提取每个季度的最后一个交易日（返回 date-only）"""
    df_dates = pd.DataFrame({"date": pd.to_datetime(dates)})
    df_dates["date_only"] = df_dates["date"].dt.date
    df_dates["year"]  = df_dates["date"].dt.year
    df_dates["quarter"] = df_dates["date"].dt.quarter
    # 对同一天的不同时间取唯一日期
    result = df_dates.groupby(["year", "quarter"])["date_only"].max().reset_index(drop=True)
    return pd.to_datetime(result)

val_qends  = get_quarter_end_dates(all_val_dates)
price_qends = get_quarter_end_dates(all_price_dates)

# 去除非标准季度末（如 2026-07-23 是近期数据，并非完整季度）
# 只保留 3/6/9/12 月末的季度
def is_quarter_end(d):
    return d.month in (3, 6, 9, 12)
val_qends  = pd.DatetimeIndex([d for d in val_qends if is_quarter_end(d)])
price_qends = pd.DatetimeIndex([d for d in price_qends if is_quarter_end(d)])

# 取交集：只有在两个文件中都有数据的季度末才有意义
common_qends = sorted(set(val_qends) & set(price_qends))
# 只保留那些"下一季度末"也在价格数据中的季度（以便计算次季度收益）
valid_qends = []
for qe in common_qends:
    nxt = price_qends[price_qends > qe]
    if len(nxt) > 0:
        valid_qends.append((qe, nxt[0]))
# 最后一个季度可能没有下个季度
valid_qends = valid_qends[:-1] if len(valid_qends) > 1 else valid_qends

print(f"可用季度: {len(valid_qends)} 个")
for qe, nxt in valid_qends:
    print(f"  {qe.strftime('%Y-%m-%d')} -> {nxt.strftime('%Y-%m-%d')}")

# ============ 4. 提取估值 & 计算次季度收益 ============
print("计算估值与收益...")
records = []

for ind_name in ind_names:
    if not ind_name or ind_name not in df_prices.columns:
        continue

    # 估值：每个季度末的 PE
    val_sub = df_val[df_val["Fetched_Ind_Name"] == ind_name].copy()
    l1 = val_sub["申万一级行业"].dropna().unique()
    l1_name = l1[0] if len(l1) > 0 else "未分类"

    prices = df_prices[ind_name].dropna()

    for qe, nxt_qe in valid_qends:
        # 季度末 PE
        pe_rows = val_sub[val_sub["tradeDate"] == qe]
        if pe_rows.empty:
            # 尝试最近的交易日
            nearby = val_sub[val_sub["tradeDate"] <= qe]
            if nearby.empty:
                continue
            pe_val = nearby.iloc[-1]["pe"]
        else:
            pe_val = pe_rows.iloc[0]["pe"]

        if pd.isna(pe_val) or pe_val <= 0:
            continue

        # 次季度收益率
        if qe not in prices.index:
            # 找最近的交易日
            nearby_p = prices[prices.index <= qe]
            if nearby_p.empty:
                continue
            p_start = nearby_p.iloc[-1]
        else:
            p_start = prices.loc[qe]

        if nxt_qe not in prices.index:
            nearby_p = prices[prices.index <= nxt_qe]
            if nearby_p.empty:
                continue
            p_end = nearby_p.iloc[-1]
        else:
            p_end = prices.loc[nxt_qe]

        if pd.isna(p_start) or pd.isna(p_end) or p_start <= 0:
            continue

        ret = (p_end / p_start) - 1

        records.append({
            "一级行业": l1_name,
            "二级行业": ind_name,
            "季度": qe.strftime("%Y%m"),
            "季度末PE": round(pe_val, 2),
            "次季度收益率": round(ret * 100, 2),  # 百分比
        })

df_result = pd.DataFrame(records)
print(f"共 {len(df_result)} 条记录, {df_result['一级行业'].nunique()} 个一级行业, {df_result['二级行业'].nunique()} 个二级行业")

# ============ 5. 输出：一级行业一个 Excel，二级行业一个 sheet ============
print("生成 Excel 文件...")

def safe_name(name):
    """清理文件名中的非法字符"""
    bad = r'\/:*?"<>|'
    for c in bad:
        name = name.replace(c, "_")
    return name[:31]  # sheet 名最大 31 字符

# 统计
total_files = 0
total_sheets = 0

for l1_name, grp_l1 in df_result.groupby("一级行业"):
    safe_l1 = safe_name(l1_name)
    xlsx_path = os.path.join(OUT_DIR, f"{safe_l1}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # 删默认 sheet

    temp_imgs = []

    for l2_name, grp_l2 in grp_l1.groupby("二级行业"):
        safe_l2 = safe_name(l2_name)

        # 画该二级行业的散点图
        fig, ax = plt.subplots(figsize=(12, 6))

        # 按季度排序
        grp_plot = grp_l2.sort_values("季度")

        # 用颜色区分不同季度
        quarter_colors = plt.cm.viridis(
            np.linspace(0.2, 0.9, len(grp_plot))
        )

        for idx, (_, row) in enumerate(grp_plot.iterrows()):
            color = quarter_colors[idx]
            ax.scatter(
                row["季度末PE"], row["次季度收益率"],
                c=[color], s=150, edgecolors="white", linewidth=0.8,
                zorder=3
            )
            # 标注季度
            ax.annotate(
                row["季度"],
                (row["季度末PE"], row["次季度收益率"]),
                textcoords="offset points",
                xytext=(8, 8),
                fontsize=9,
                fontweight="bold",
                color="black",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                          edgecolor=color, alpha=0.85),
            )

        ax.set_xlabel("季度末 PE (市盈率)", fontsize=11)
        ax.set_ylabel("次季度收益率 (%)", fontsize=11)
        ax.set_title(
            f"{l1_name} · {l2_name} — 季度末 PE vs 次季度收益率",
            fontsize=13, fontweight="bold"
        )
        ax.axhline(y=0, color="gray", linestyle="--", linewidth=0.8, alpha=0.5)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(left=0)

        plt.tight_layout()

        # 保存到临时图片
        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()
        fig.savefig(tmp.name, dpi=150, bbox_inches="tight")
        plt.close(fig)
        temp_imgs.append(tmp.name)

        # 创建 sheet 并插入图片
        ws = wb.create_sheet(title=safe_l2)
        img = OpxlImage(tmp.name)
        img.width = 840
        img.height = 420
        ws.add_image(img, "A1")

        # 右侧数据表
        data_start_col = 16
        headers = ["季度(YYYYMM)", "季度末PE", "次季度收益率(%)"]
        for j, h in enumerate(headers):
            ws.cell(row=1, column=data_start_col + j, value=h)

        for i, (_, row) in enumerate(grp_plot.iterrows()):
            ws.cell(row=i + 2, column=data_start_col, value=row["季度"])
            ws.cell(row=i + 2, column=data_start_col + 1, value=row["季度末PE"])
            ws.cell(row=i + 2, column=data_start_col + 2, value=row["次季度收益率"])

        for j in range(3):
            ws.column_dimensions[get_column_letter(data_start_col + j)].width = 18

        total_sheets += 1

    wb.save(xlsx_path)
    wb.close()

    # 清理临时图片
    for tmp_path in temp_imgs:
        try:
            os.remove(tmp_path)
        except:
            pass

    total_files += 1
    print(f"  [OK] {safe_l1}.xlsx: {grp_l1['二级行业'].nunique()} 个二级行业")

# 保存汇总 CSV
df_result.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

print(f"\n完成! {total_files} 个 Excel 文件, {total_sheets} 个 sheet")
print(f"输出目录: {OUT_DIR}")
print(f"数据 CSV: {OUT_CSV}")
