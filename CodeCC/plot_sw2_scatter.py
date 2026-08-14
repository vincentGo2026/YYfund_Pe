import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

SRC = Path(r'D:\CC\DB\data\Fund_IRG_updated_v2.xlsx')
OUT = Path(r'D:\CC\Mid\估值')

# 读取数据
wb = load_workbook(SRC, data_only=True)
ws = wb['申万二级行业分布']

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    manager = r[0]
    pct = r[3]
    if manager == '实盘主动' and pct is not None and pct > 0:
        rows.append(r)

rows.sort(key=lambda r: r[3] if r[3] else 0, reverse=True)

# 准备散点数据 (仅保留PB和PE百分位都有的行)
scatter_data = []
for r in rows:
    ind1, ind2, pct = r[1], r[2], r[3]
    pb_pct, pe_pct = r[5], r[7]  # PB百分位, PE百分位
    method = r[8]
    if pb_pct is not None and pe_pct is not None:
        label = str(ind2)
        scatter_data.append({
            'label': label,
            'ind1': str(ind1),
            'pb_pct': pb_pct,
            'pe_pct': pe_pct,
            'weight': pct,
            'method': method,
        })

# 计算气泡大小 (市值占比越大,气泡越大)
max_w = max(d['weight'] for d in scatter_data) if scatter_data else 1
min_size, max_size = 30, 1200

valid_data = [d for d in scatter_data if d['pb_pct'] is not None and d['pe_pct'] is not None]

fig, ax = plt.subplots(figsize=(16, 12))

for d in valid_data:
    size = min_size + (d['weight'] / max_w) * (max_size - min_size)
    color = 'steelblue' if d['method'] == 'PB' else 'darkorange'
    ax.scatter(d['pb_pct'], d['pe_pct'], s=size, c=color, alpha=0.7, edgecolors='white', linewidth=1)
    offset = 2 if d['pb_pct'] < 50 else -2
    ax.annotate(d['label'], (d['pb_pct'], d['pe_pct']),
                textcoords="offset points", xytext=(offset, 5),
                fontsize=7, alpha=0.85,
                arrowprops=dict(arrowstyle='->', color='gray', lw=0.5, alpha=0.5))

# 象限分割线
ax.axhline(y=50, color='gray', linestyle='--', alpha=0.4)
ax.axvline(x=50, color='gray', linestyle='--', alpha=0.4)

# 象限标注
ax.text(10, 95, 'PB低估\nPE高估', fontsize=11, color='gray', ha='center', alpha=0.7)
ax.text(90, 95, 'PB高估\nPE高估', fontsize=11, color='gray', ha='center', alpha=0.7)
ax.text(10, 5, 'PB低估\nPE低估', fontsize=11, color='gray', ha='center', alpha=0.7)
ax.text(90, 5, 'PB高估\nPE低估', fontsize=11, color='gray', ha='center', alpha=0.7)

ax.set_xlabel('PB历史百分位(%)', fontsize=14)
ax.set_ylabel('PE历史百分位(%)', fontsize=14)
ax.set_title('实盘主动 申万二级行业 PB vs PE 历史百分位散点图\n(气泡大小=市值占比, 蓝色=PB估值行业, 橙色=PE估值行业)',
             fontsize=15, fontweight='bold')
ax.set_xlim(-5, 105)
ax.set_ylim(-5, 105)
ax.grid(True, alpha=0.3)

# 高占比行业图例
from matplotlib.lines import Line2D
legend_elements = [
    Line2D([0], [0], marker='o', color='w', markerfacecolor='steelblue', markersize=12, label='PB估值行业'),
    Line2D([0], [0], marker='o', color='w', markerfacecolor='darkorange', markersize=12, label='PE估值行业'),
]
ax.legend(handles=legend_elements, loc='upper right', fontsize=11)

fig.tight_layout()

out_path = OUT / 'sw2_valuation_scatter.png'
fig.savefig(out_path, dpi=150, bbox_inches='tight')
plt.close()
print(f'[OK] {out_path}')
print(f'共 {len(valid_data)} 个有效数据点')
