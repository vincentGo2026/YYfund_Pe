# -*- coding: utf-8 -*-
r"""
季度末 PB 历史百分位 vs 次季度收益率 — 一级行业一个 Excel，每个二级行业一个 sheet
数据源：
  - D:\CC\DB\data\datayes_all_SW_Industries_Level2_mapped.csv  (估值 PB，R列)
  - D:\CC\DB\data\sw2_market_data_20210501_20260724.csv         (行情价格，纵向拼接)
计算逻辑：
  - 从 2022/6/30 开始，每个季度末计算当前 PB 在其自身历史（从最早数据到当前日）中的百分位
  - PB百分位 = 历史PB小于等于当前PB的天数 / 历史总天数 × 100
输出：
  - D:\CC\Mid\估值\REStemp\pb_pct\  每个一级行业一个 .xlsx 文件
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpxlImage
from openpyxl.utils import get_column_letter
import os
import tempfile

# ============ 路径 ============
VAL_PATH   = r"D:\CC\DB\data\datayes_all_SW_Industries_Level2_mapped.csv"
PRICE_PATH = r"D:\CC\DB\data\sw2_market_data_20210501_20260724.csv"
OUT_DIR    = r"D:\CC\Mid\估值\REStemp\pb_pct"
OUT_CSV    = r"D:\CC\Mid\估值\REStemp\pb_pct\quarterly_pb_pct_vs_return.csv"
START_DATE  = "2022-06-30"  # 百分位统计起始日

os.makedirs(OUT_DIR, exist_ok=True)

# 中文字体
plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

# ============ 1. 读取估值数据 ============
print("读取估值数据...")
df_val = pd.read_csv(VAL_PATH, parse_dates=["tradeDate"])
df_val["pb"] = pd.to_numeric(df_val["pb"], errors="coerce")
df_val = df_val.dropna(subset=["pb"])
df_val = df_val[df_val["pb"] > 0]
df_val["tradeDate"] = pd.to_datetime(df_val["tradeDate"].dt.date)
df_val = df_val.sort_values(["Fetched_Ind_Name", "tradeDate"])
print(f"  估值记录: {len(df_val)} 条, 日期范围: {df_val['tradeDate'].min().strftime('%Y-%m-%d')} ~ {df_val['tradeDate'].max().strftime('%Y-%m-%d')}")

# ============ 2. 读取价格数据（长格式） ============
print("读取价格数据...")
df_price_raw = pd.read_csv(PRICE_PATH, encoding="utf-8-sig")
df_price_raw["tradeDate"] = pd.to_datetime(df_price_raw["tradeDate"])

df_prices = df_price_raw.pivot_table(
    index="tradeDate",
    columns="申万二级行业",
    values="closeIndex",
    aggfunc="last"
)
df_prices.index = pd.to_datetime(df_prices.index.date)
df_prices = df_prices.sort_index()

print(f"  行业数: {df_prices.shape[1]}, 交易日: {df_prices.shape[0]}")

# ============ 3. 确定季度末日期 ============
print("确定季度区间...")
all_val_dates = sorted(df_val["tradeDate"].unique())
all_price_dates = sorted(df_prices.index)


def get_quarter_end_dates(dates):
    df_dates = pd.DataFrame({"date": pd.to_datetime(dates)})
    df_dates["date_only"] = df_dates["date"].dt.date
    df_dates["year"] = df_dates["date"].dt.year
    df_dates["quarter"] = df_dates["date"].dt.quarter
    result = df_dates.groupby(["year", "quarter"])["date_only"].max().reset_index(drop=True)
    return pd.to_datetime(result)


val_qends = get_quarter_end_dates(all_val_dates)
price_qends = get_quarter_end_dates(all_price_dates)


def is_quarter_end(d):
    return d.month in (3, 6, 9, 12)


val_qends = pd.DatetimeIndex([d for d in val_qends if is_quarter_end(d)])
price_qends = pd.DatetimeIndex([d for d in price_qends if is_quarter_end(d)])

# 取交集，且从 START_DATE 开始
common_qends = sorted(set(val_qends) & set(price_qends))
start_dt = pd.Timestamp(START_DATE)
common_qends = [d for d in common_qends if d >= start_dt]

valid_qends = []
for qe in common_qends:
    nxt = price_qends[price_qends > qe]
    if len(nxt) > 0:
        valid_qends.append((qe, nxt[0]))

print(f"可用季度: {len(valid_qends)} 个 (>= {START_DATE})")
for qe, nxt in valid_qends:
    print(f"  {qe.strftime('%Y-%m-%d')} -> {nxt.strftime('%Y-%m-%d')}")

# ============ 4. 计算 PB 历史百分位 + 次季度收益 ============
print("计算 PB 历史百分位与收益...")
print("  (对每个行业，百分位 = 当前 PB 在自身历史中的排位 / 历史总天数 × 100)")

records = []
price_industries = list(df_prices.columns)
start_dt = pd.Timestamp(START_DATE)

for ind_name in price_industries:
    # 估值匹配
    val_name = ind_name + "(申万)"
    val_sub = df_val[df_val["Fetched_Ind_Name"] == val_name].copy()
    if val_sub.empty:
        val_sub = df_val[df_val["Fetched_Ind_Name"] == ind_name].copy()
    if val_sub.empty:
        continue
    val_sub = val_sub.sort_values("tradeDate").reset_index(drop=True)

    l1 = val_sub["申万一级行业"].dropna().unique()
    l1_name = l1[0] if len(l1) > 0 else "未分类"

    prices = df_prices[ind_name].dropna()

    for qe, nxt_qe in valid_qends:
        # 季度末 PB：取 qe 当天或之前最近的估值
        pb_rows = val_sub[val_sub["tradeDate"] == qe]
        if pb_rows.empty:
            nearby = val_sub[val_sub["tradeDate"] <= qe]
            if nearby.empty:
                continue
            pb_val = nearby.iloc[-1]["pb"]
            pb_idx = nearby.index[-1]
        else:
            pb_val = pb_rows.iloc[0]["pb"]
            pb_idx = pb_rows.index[0]

        if pd.isna(pb_val) or pb_val <= 0:
            continue

        # 计算历史百分位：从最早到当前日
        hist_pb = val_sub.loc[:pb_idx, "pb"]
        if len(hist_pb) < 10:  # 历史数据太少，不够可靠
            continue

        pct = (hist_pb <= pb_val).sum() / len(hist_pb) * 100

        # 次季度收益率
        if qe not in prices.index:
            nearby_p = prices[prices.index <= qe]
            if nearby_p.empty:
                continue
            p_start = nearby_p.iloc[-1]
        else:
            p_start = prices.loc[qe]

        if nxt_qe not in prices.index:
            nearby_p = prices[prices.index <= nxt_qe]
            if nearby_p.empty:
                continue
            p_end = nearby_p.iloc[-1]
        else:
            p_end = prices.loc[nxt_qe]

        if pd.isna(p_start) or pd.isna(p_end) or p_start <= 0:
            continue

        ret = (p_end / p_start) - 1

        records.append({
            "一级行业": l1_name,
            "二级行业": ind_name,
            "季度": qe.strftime("%Y%m"),
            "PB历史百分位": round(pct, 2),
            "季度末PB": round(pb_val, 2),
            "次季度收益率": round(ret * 100, 2),
        })

df_result = pd.DataFrame(records)
print(f"共 {len(df_result)} 条记录, {df_result['一级行业'].nunique()} 个一级行业, {df_result['二级行业'].nunique()} 个二级行业")

# 打印百分位分布
print(f"\nPB 历史百分位分布:")
print(f"  mean={df_result['PB历史百分位'].mean():.1f}%")
print(f"  median={df_result['PB历史百分位'].median():.1f}%")
print(f"  min={df_result['PB历史百分位'].min():.1f}%")
print(f"  max={df_result['PB历史百分位'].max():.1f}%")
print()

# ============ 5. 输出：一级行业一个 Excel，二级行业一个 sheet ============
print("生成 Excel 文件...")


def safe_name(name):
    bad = r'\/:*?"<>|'
    for c in bad:
        name = name.replace(c, "_")
    return name[:31]


total_files = 0
total_sheets = 0

for l1_name, grp_l1 in df_result.groupby("一级行业"):
    safe_l1 = safe_name(l1_name)
    xlsx_path = os.path.join(OUT_DIR, f"{safe_l1}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    temp_imgs = []

    for l2_name, grp_l2 in grp_l1.groupby("二级行业"):
        safe_l2 = safe_name(l2_name)

        fig, ax = plt.subplots(figsize=(12, 6))
        grp_plot = grp_l2.sort_values("季度").reset_index(drop=True)

        # 蓝 → 橙渐变：最早蓝色，最新橙色
        n = max(len(grp_plot), 1)
        blue_orange_cmap = LinearSegmentedColormap.from_list(
            "blue_orange", ["#0073cc", "#f08c0a"], N=n
        )
        quarter_colors = [blue_orange_cmap(i / max(n - 1, 1)) for i in range(n)]

        # 画连线（按时间顺序）
        x_vals = grp_plot["PB历史百分位"].values
        y_vals = grp_plot["次季度收益率"].values
        for i in range(len(x_vals) - 1):
            ax.plot(
                [x_vals[i], x_vals[i + 1]],
                [y_vals[i], y_vals[i + 1]],
                color="gray", linewidth=0.8, alpha=0.5, zorder=2
            )

        for idx, (_, row) in enumerate(grp_plot.iterrows()):
            color = quarter_colors[idx]
            ax.scatter(
                row["PB历史百分位"], row["次季度收益率"],
                c=[color], s=150, edgecolors="white", linewidth=0.8,
                zorder=4
            )
            ax.annotate(
                row["季度"],
                (row["PB历史百分位"], row["次季度收益率"]),
                textcoords="offset points",
                xytext=(8, 8),
                fontsize=9,
                fontweight="bold",
                color="black",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                          edgecolor=color, alpha=0.85),
            )

        ax.set_xlabel("季度末 PB 历史百分位 (%)", fontsize=11)
        ax.set_ylabel("次季度收益率 (%)", fontsize=11)
        ax.set_title(
            f"{l1_name} · {l2_name} — 季度末 PB 历史百分位 vs 次季度收益率",
            fontsize=13, fontweight="bold"
        )
        ax.axhline(y=0, color="gray", linestyle="--", linewidth=0.8, alpha=0.5)
        ax.axvline(x=50, color="red", linestyle="--", linewidth=0.6, alpha=0.3)
        ax.grid(True, alpha=0.3)

        # 高百分位=贵，右；低百分位=便宜，左
        ax.set_xlim(0, 100)
        # 分区标注
        ax.axvspan(0, 20, alpha=0.05, color='green', label='极度低估')
        ax.axvspan(20, 40, alpha=0.03, color='lightgreen', label='低估')
        ax.axvspan(60, 80, alpha=0.03, color='lightcoral', label='高估')
        ax.axvspan(80, 100, alpha=0.05, color='red', label='极度高估')
        ax.legend(loc='upper right', fontsize=8)

        plt.tight_layout()

        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()
        fig.savefig(tmp.name, dpi=150, bbox_inches="tight")
        plt.close(fig)
        temp_imgs.append(tmp.name)

        ws = wb.create_sheet(title=safe_l2)
        img = OpxlImage(tmp.name)
        img.width = 840
        img.height = 420
        ws.add_image(img, "A1")

        data_start_col = 16
        headers = ["季度(YYYYMM)", "PB历史百分位(%)", "季度末PB", "次季度收益率(%)"]
        for j, h in enumerate(headers):
            ws.cell(row=1, column=data_start_col + j, value=h)

        for i, (_, row) in enumerate(grp_plot.iterrows()):
            ws.cell(row=i + 2, column=data_start_col, value=row["季度"])
            ws.cell(row=i + 2, column=data_start_col + 1, value=row["PB历史百分位"])
            ws.cell(row=i + 2, column=data_start_col + 2, value=row["季度末PB"])
            ws.cell(row=i + 2, column=data_start_col + 3, value=row["次季度收益率"])

        for j in range(4):
            ws.column_dimensions[get_column_letter(data_start_col + j)].width = 18

        total_sheets += 1

    wb.save(xlsx_path)
    wb.close()

    for tmp_path in temp_imgs:
        try:
            os.remove(tmp_path)
        except:
            pass

    total_files += 1
    print(f"  [OK] {safe_l1}.xlsx: {grp_l1['二级行业'].nunique()} 个二级行业")

# 保存汇总 CSV
df_result.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

print(f"\n完成! {total_files} 个 Excel 文件, {total_sheets} 个 sheet")
print(f"输出目录: {OUT_DIR}")
print(f"数据 CSV: {OUT_CSV}")
