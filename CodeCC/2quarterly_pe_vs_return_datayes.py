# -*- coding: utf-8 -*-
r"""
季度估值 vs 次季度收益率 — 一级行业一个 Excel，每个二级行业一个 sheet
数据源：
  - D:\CC\DB\data\datayes_all_SW_Industries_Level2_mapped.csv  (估值 PE)
  - D:\CC\DB\data\sw2_market_data_20210501_20260724.csv         (行情价格，纵向拼接)
输出：
  - output/pe_vs_return_v2/  每个一级行业一个 .xlsx 文件，每个二级行业一个 sheet
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpxlImage
from openpyxl.utils import get_column_letter
import os
import tempfile

# ============ 路径 ============
VAL_PATH   = r"D:\CC\DB\data\datayes_all_SW_Industries_Level2_mapped.csv"
PRICE_PATH = r"D:\CC\DB\data\sw2_market_data_20210501_20260724.csv"
OUT_DIR    = r"D:\CC\Mid\估值\REStemp"
OUT_CSV    = r"D:\CC\Mid\估值\REStemp\quarterly_pe_vs_return_data_v2.csv"

os.makedirs(OUT_DIR, exist_ok=True)

# 中文字体
plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

# ============ 1. 读取估值数据 ============
print("读取估值数据...")
df_val = pd.read_csv(VAL_PATH, parse_dates=["tradeDate"])
df_val = df_val[["tradeDate", "pe", "Fetched_Ind_Name", "申万一级行业"]].copy()
df_val = df_val.dropna(subset=["pe"])
df_val["pe"] = pd.to_numeric(df_val["pe"], errors="coerce")
df_val = df_val[df_val["pe"] > 0]
# 统一为 date-only
df_val["tradeDate"] = df_val["tradeDate"].dt.date
df_val["tradeDate"] = pd.to_datetime(df_val["tradeDate"])

# ============ 2. 读取价格数据（长格式） ============
print("读取价格数据...")
df_price_raw = pd.read_csv(PRICE_PATH, encoding="utf-8-sig")
df_price_raw["tradeDate"] = pd.to_datetime(df_price_raw["tradeDate"])

# 长格式：每个行业的时间序列纵向拼接
# 列：tradeDate, closeIndex, 申万二级行业
# 转换为宽格式便于操作
df_prices = df_price_raw.pivot_table(
    index="tradeDate",
    columns="申万二级行业",
    values="closeIndex",
    aggfunc="last"
)
df_prices.index = pd.to_datetime(df_prices.index.date)  # date-only
df_prices = df_prices.sort_index()

print(f"  行业数: {df_prices.shape[1]}, 交易日: {df_prices.shape[0]}")
print(f"  日期范围: {df_prices.index.min().strftime('%Y-%m-%d')} ~ {df_prices.index.max().strftime('%Y-%m-%d')}")

# ============ 3. 确定季度末日期 ============
print("确定季度区间...")
all_val_dates = sorted(df_val["tradeDate"].unique())
all_price_dates = sorted(df_prices.index)


def get_quarter_end_dates(dates):
    """从日期列表中提取每个季度的最后一个交易日"""
    df_dates = pd.DataFrame({"date": pd.to_datetime(dates)})
    df_dates["date_only"] = df_dates["date"].dt.date
    df_dates["year"] = df_dates["date"].dt.year
    df_dates["quarter"] = df_dates["date"].dt.quarter
    result = df_dates.groupby(["year", "quarter"])["date_only"].max().reset_index(drop=True)
    return pd.to_datetime(result)


val_qends = get_quarter_end_dates(all_val_dates)
price_qends = get_quarter_end_dates(all_price_dates)


def is_quarter_end(d):
    return d.month in (3, 6, 9, 12)


val_qends = pd.DatetimeIndex([d for d in val_qends if is_quarter_end(d)])
price_qends = pd.DatetimeIndex([d for d in price_qends if is_quarter_end(d)])

# 取交集
common_qends = sorted(set(val_qends) & set(price_qends))
valid_qends = []
for qe in common_qends:
    nxt = price_qends[price_qends > qe]
    if len(nxt) > 0:
        valid_qends.append((qe, nxt[0]))
# 最后一个季度可能没有下个季度
valid_qends = valid_qends[:-1] if len(valid_qends) > 1 else valid_qends

print(f"可用季度: {len(valid_qends)} 个")
for qe, nxt in valid_qends:
    print(f"  {qe.strftime('%Y-%m-%d')} -> {nxt.strftime('%Y-%m-%d')}")

# ============ 4. 提取估值 & 计算次季度收益 ============
print("计算估值与收益...")

# 估值行业名：半导体(申万) → 去掉 (申万) 匹配价格行业名
df_val["ind_short"] = df_val["Fetched_Ind_Name"].str.replace("(申万)", "", regex=False)

records = []

price_industries = list(df_prices.columns)

for ind_name in price_industries:
    # 估值：每个季度末的 PE（估值名需加 (申万) 后缀匹配）
    val_name = ind_name + "(申万)"
    val_sub = df_val[df_val["Fetched_Ind_Name"] == val_name].copy()
    if val_sub.empty:
        # 尝试直接用原名
        val_sub = df_val[df_val["Fetched_Ind_Name"] == ind_name].copy()
    if val_sub.empty:
        continue

    l1 = val_sub["申万一级行业"].dropna().unique()
    l1_name = l1[0] if len(l1) > 0 else "未分类"

    prices = df_prices[ind_name].dropna()

    for qe, nxt_qe in valid_qends:
        # 季度末 PE
        pe_rows = val_sub[val_sub["tradeDate"] == qe]
        if pe_rows.empty:
            nearby = val_sub[val_sub["tradeDate"] <= qe]
            if nearby.empty:
                continue
            pe_val = nearby.iloc[-1]["pe"]
        else:
            pe_val = pe_rows.iloc[0]["pe"]

        if pd.isna(pe_val) or pe_val <= 0:
            continue

        # 次季度收益率
        if qe not in prices.index:
            nearby_p = prices[prices.index <= qe]
            if nearby_p.empty:
                continue
            p_start = nearby_p.iloc[-1]
        else:
            p_start = prices.loc[qe]

        if nxt_qe not in prices.index:
            nearby_p = prices[prices.index <= nxt_qe]
            if nearby_p.empty:
                continue
            p_end = nearby_p.iloc[-1]
        else:
            p_end = prices.loc[nxt_qe]

        if pd.isna(p_start) or pd.isna(p_end) or p_start <= 0:
            continue

        ret = (p_end / p_start) - 1

        records.append({
            "一级行业": l1_name,
            "二级行业": ind_name,
            "季度": qe.strftime("%Y%m"),
            "季度末PE": round(pe_val, 2),
            "次季度收益率": round(ret * 100, 2),
        })

df_result = pd.DataFrame(records)
print(f"共 {len(df_result)} 条记录, {df_result['一级行业'].nunique()} 个一级行业, {df_result['二级行业'].nunique()} 个二级行业")

# ============ 5. 输出：一级行业一个 Excel，二级行业一个 sheet ============
print("生成 Excel 文件...")


def safe_name(name):
    bad = r'\/:*?"<>|'
    for c in bad:
        name = name.replace(c, "_")
    return name[:31]


total_files = 0
total_sheets = 0

for l1_name, grp_l1 in df_result.groupby("一级行业"):
    safe_l1 = safe_name(l1_name)
    xlsx_path = os.path.join(OUT_DIR, f"{safe_l1}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    temp_imgs = []

    for l2_name, grp_l2 in grp_l1.groupby("二级行业"):
        safe_l2 = safe_name(l2_name)

        fig, ax = plt.subplots(figsize=(12, 6))
        grp_plot = grp_l2.sort_values("季度").reset_index(drop=True)

        # 蓝 → 橙渐变：最早蓝色(0,0.45,0.8)，最新橙色(0.95,0.55,0.1)
        n = max(len(grp_plot), 1)
        from matplotlib.colors import LinearSegmentedColormap
        blue_orange_cmap = LinearSegmentedColormap.from_list(
            "blue_orange", ["#0073cc", "#f08c0a"], N=n
        )
        quarter_colors = [blue_orange_cmap(i / max(n - 1, 1)) for i in range(n)]

        # 画连线（按时间顺序）
        x_vals = grp_plot["季度末PE"].values
        y_vals = grp_plot["次季度收益率"].values
        for i in range(len(x_vals) - 1):
            ax.plot(
                [x_vals[i], x_vals[i + 1]],
                [y_vals[i], y_vals[i + 1]],
                color="gray", linewidth=0.8, alpha=0.5, zorder=2
            )

        for idx, (_, row) in enumerate(grp_plot.iterrows()):
            color = quarter_colors[idx]
            ax.scatter(
                row["季度末PE"], row["次季度收益率"],
                c=[color], s=150, edgecolors="white", linewidth=0.8,
                zorder=4
            )
            ax.annotate(
                row["季度"],
                (row["季度末PE"], row["次季度收益率"]),
                textcoords="offset points",
                xytext=(8, 8),
                fontsize=9,
                fontweight="bold",
                color="black",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                          edgecolor=color, alpha=0.85),
            )

        ax.set_xlabel("季度末 PE (市盈率)", fontsize=11)
        ax.set_ylabel("次季度收益率 (%)", fontsize=11)
        ax.set_title(
            f"{l1_name} · {l2_name} — 季度末 PE vs 次季度收益率",
            fontsize=13, fontweight="bold"
        )
        ax.axhline(y=0, color="gray", linestyle="--", linewidth=0.8, alpha=0.5)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(left=0)

        plt.tight_layout()

        tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        tmp.close()
        fig.savefig(tmp.name, dpi=150, bbox_inches="tight")
        plt.close(fig)
        temp_imgs.append(tmp.name)

        ws = wb.create_sheet(title=safe_l2)
        img = OpxlImage(tmp.name)
        img.width = 840
        img.height = 420
        ws.add_image(img, "A1")

        data_start_col = 16
        headers = ["季度(YYYYMM)", "季度末PE", "次季度收益率(%)"]
        for j, h in enumerate(headers):
            ws.cell(row=1, column=data_start_col + j, value=h)

        for i, (_, row) in enumerate(grp_plot.iterrows()):
            ws.cell(row=i + 2, column=data_start_col, value=row["季度"])
            ws.cell(row=i + 2, column=data_start_col + 1, value=row["季度末PE"])
            ws.cell(row=i + 2, column=data_start_col + 2, value=row["次季度收益率"])

        for j in range(3):
            ws.column_dimensions[get_column_letter(data_start_col + j)].width = 18

        total_sheets += 1

    wb.save(xlsx_path)
    wb.close()

    for tmp_path in temp_imgs:
        try:
            os.remove(tmp_path)
        except:
            pass

    total_files += 1
    print(f"  [OK] {safe_l1}.xlsx: {grp_l1['二级行业'].nunique()} 个二级行业")

# 保存汇总 CSV
df_result.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

print(f"\n完成! {total_files} 个 Excel 文件, {total_sheets} 个 sheet")
print(f"输出目录: {OUT_DIR}")
print(f"数据 CSV: {OUT_CSV}")
