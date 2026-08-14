"""对实盘主动_申万二级行业估值 使用与 行业投资价值综合评估_混合版 同样的评估格式"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC_ASSESS = r"D:\CC\Mid\估值\申万\行业投资价值综合评估_混合版.xlsx"
SRC_SHIPAN = r"D:\CC\Mid\估值\实盘主动_申万二级行业估值.xlsx"
OUT = r"D:\CC\Mid\估值\实盘主动_申万二级行业评估.xlsx"
OUT_DETAIL = r"D:\CC\Mid\估值\实盘主动_申万二级行业估值明细.xlsx"

# ———— 1. 读取混合评估的 一级行业 映射 ————
wb_assess = openpyxl.load_workbook(SRC_ASSESS)
ws_a = wb_assess["混合评估"]
l1_map = {}
for row in ws_a.iter_rows(min_row=2, values_only=True):
    name = row[0]  # 行业
    l1_map[name] = {
        "投资评级": row[1],
        "综合评分": row[2],
        "估值评分": row[3],
        "动量评分": row[4],
        "风险等级": row[5],
        "估值状态": row[6],
        "估值依据": row[7],
        "最新PE": row[8],
        "PE历史分位数(%)": row[9],
        "PE状态": row[10],
        "最新PB": row[11],
        "PB历史分位数(%)": row[12],
        "PB状态": row[13],
        "动量状态": row[14],
        "2026年收益(%)": row[15],
        "近3年累计收益(%)": row[16],
        "年化波动率(%)": row[17],
        "年胜率(%)": row[18],
        "收益风险比": row[19],
    }

# ———— 2. 读取实盘主动持仓 ————
wb_shipan = openpyxl.load_workbook(SRC_SHIPAN)
ws_s = wb_shipan.active
shipan_rows = []
for row in ws_s.iter_rows(min_row=2, values_only=True):
    shipan_rows.append({
        "申万一级行业": row[0],
        "申万二级行业": row[1],
        "市值占比%": row[2],
        "最新PB": row[3],
        "PB历史百分位(%)": row[4],
        "最新PE": row[5],
        "PE历史百分位(%)": row[6],
    })

# ———— 3. PB估值行业规则 ————
pb_industries = {"银行", "非银金融", "电力设备", "交通运输", "有色金属", "建筑装饰", "食品饮料", "家用电器", "商贸零售"}

# ———— 4. 百分位 → 状态 映射 ————
def pct_to_state(pct):
    """将百分位映射为估值状态"""
    if pct is None:
        return None
    if pct < 20:
        return "深度低估"
    elif pct < 40:
        return "低估"
    elif pct < 60:
        return "合理"
    elif pct < 80:
        return "偏高"
    else:
        return "高估"

def state_to_score(state):
    """估值状态 → 估值评分"""
    mapping = {"深度低估": 3, "低估": 2, "合理": 2, "偏高": 1, "高估": 0}
    return mapping.get(state, None)

# ———— 5. 组装输出 ————
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "实盘主动评估"

headers = [
    "申万一级行业", "申万二级行业", "市值占比%",
    "投资评级", "综合评分", "估值评分", "动量评分", "风险等级",
    "估值状态", "估值依据",
    "最新PE", "PE历史分位数(%)", "PE状态",
    "最新PB", "PB历史分位数(%)", "PB状态",
    "动量状态",
    "2026年收益(%)", "近3年累计收益(%)", "年化波动率(%)", "年胜率(%)", "收益风险比",
]
out_ws.append(headers)

for r in shipan_rows:
    l1 = r["申万一级行业"]
    l2 = r["申万二级行业"]
    pct = r["市值占比%"]
    pb_val = r["最新PB"]
    pb_pct = r["PB历史百分位(%)"]
    pe_val = r["最新PE"]
    pe_pct = r["PE历史百分位(%)"]

    # 估值依据
    basis = "PB" if l1 in pb_industries else "PE"

    # PE/PB 状态
    pe_state = pct_to_state(pe_pct)
    pb_state = pct_to_state(pb_pct)

    # 估值状态（使用对应估值依据的状态）
    val_state = pb_state if basis == "PB" else pe_state

    # 估值评分
    val_score = state_to_score(val_state)

    # 从一级行业映射动量/风险/收益等
    l1_info = l1_map.get(l1, {})

    out_ws.append([
        l1,
        l2,
        pct,
        l1_info.get("投资评级"),
        l1_info.get("综合评分"),
        val_score,
        l1_info.get("动量评分"),
        l1_info.get("风险等级"),
        val_state,
        basis,
        pe_val,
        pe_pct,
        pe_state,
        pb_val,
        pb_pct,
        pb_state,
        l1_info.get("动量状态"),
        l1_info.get("2026年收益(%)"),
        l1_info.get("近3年累计收益(%)"),
        l1_info.get("年化波动率(%)"),
        l1_info.get("年胜率(%)"),
        l1_info.get("收益风险比"),
    ])

# ———— 6. 格式化 ————
header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

for cell in out_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 评级着色
rating_colors = {
    "积极配置": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "标配": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "谨慎观望": PatternFill(start_color="F4B4C2", end_color="F4B4C2", fill_type="solid"),
    "规避": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
}

score_colors = {
    3: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    2: PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    1: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    0: PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
}

state_colors = {
    "深度低估": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
    "低估": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
    "合理": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "偏高": PatternFill(start_color="F4A460", end_color="F4A460", fill_type="solid"),
    "高估": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
}

risk_colors = {
    "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "高": PatternFill(start_color="F4B4C2", end_color="F4B4C2", fill_type="solid"),
}

for row in out_ws.iter_rows(min_row=2, max_col=len(headers)):
    # 投资评级 (col D, index 3)
    rating = row[3].value
    if rating in rating_colors:
        row[3].fill = rating_colors[rating]

    # 估值评分 (col F, index 5)
    vs = row[5].value
    if vs in score_colors:
        row[5].fill = score_colors[vs]

    # 风险等级 (col H, index 7)
    risk = row[7].value
    if risk in risk_colors:
        row[7].fill = risk_colors[risk]

    # 估值状态 (col I, index 8)
    vs2 = row[8].value
    if vs2 in state_colors:
        row[8].fill = state_colors[vs2]

    # 估值依据 (col J, index 9): PB=浅蓝, PE=浅绿
    if row[9].value == "PB":
        row[9].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    elif row[9].value == "PE":
        row[9].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # PE状态 (col M, index 12)
    if row[12].value in state_colors:
        row[12].fill = state_colors[row[12].value]

    # PB状态 (col P, index 15)
    if row[15].value in state_colors:
        row[15].fill = state_colors[row[15].value]

    # 边框和对齐
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=9)

    # 申万二级行业 左对齐
    row[1].alignment = Alignment(horizontal="left", vertical="center")

# 列宽
col_widths = {
    "A": 12, "B": 16, "C": 10,
    "D": 12, "E": 12, "F": 12, "G": 12, "H": 12,
    "I": 12, "J": 12,
    "K": 14, "L": 18, "M": 14,
    "N": 14, "O": 18, "P": 14,
    "Q": 12,
    "R": 16, "S": 18, "T": 18, "U": 14, "V": 14,
}
for col_letter, width in col_widths.items():
    out_ws.column_dimensions[col_letter].width = width

# 冻结首行
out_ws.freeze_panes = "A2"

out_wb.save(OUT)
print(f"已生成 -> {OUT}")
print(f"共 {len(shipan_rows)} 行（{len(shipan_rows)-1} 个申万二级行业）")

# 统计摘要
print("\n【估值状态分布】")
from collections import Counter
state_dist = Counter(r2[8] for r2 in out_ws.iter_rows(min_row=2, values_only=True))
for s, c in state_dist.most_common():
    print(f"  {s}: {c} 个")

print("\n【估值依据分布】")
basis_dist = Counter(r2[9] for r2 in out_ws.iter_rows(min_row=2, values_only=True))
for b, c in basis_dist.most_common():
    print(f"  {b}: {c} 个")

# ———— 7. 生成估值明细（精简版，10列） ————
detail_wb = openpyxl.Workbook()
detail_ws = detail_wb.active
detail_ws.title = "持仓估值"

detail_headers = [
    "申万一级行业", "申万二级行业", "市值占比%",
    "估值依据",
    "最新PE", "PE历史分位数(%)", "PE状态",
    "最新PB", "PB历史分位数(%)", "PB状态",
]
detail_ws.append(detail_headers)

detail_rows = []
for r in shipan_rows:
    l1 = r["申万一级行业"]
    l2 = r["申万二级行业"]
    pct = r["市值占比%"]
    pb_val = r["最新PB"]
    pb_pct = r["PB历史百分位(%)"]
    pe_val = r["最新PE"]
    pe_pct = r["PE历史百分位(%)"]
    basis = "PB" if l1 in pb_industries else "PE"
    pe_state = pct_to_state(pe_pct)
    pb_state = pct_to_state(pb_pct)
    detail_rows.append([l1, l2, pct, basis, pe_val, pe_pct, pe_state, pb_val, pb_pct, pb_state])

detail_rows.sort(key=lambda r: r[2] if r[2] else 0, reverse=True)
for dr in detail_rows:
    detail_ws.append(dr)

# 格式化
for cell in detail_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row in detail_ws.iter_rows(min_row=2, max_col=len(detail_headers)):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=9)
    row[1].alignment = Alignment(horizontal="left", vertical="center")
    # 估值依据着色
    if row[3].value == "PB":
        row[3].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    elif row[3].value == "PE":
        row[3].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    # PE/PB 状态着色
    if row[6].value in state_colors:
        row[6].fill = state_colors[row[6].value]
    if row[9].value in state_colors:
        row[9].fill = state_colors[row[9].value]

detail_widths = {"A": 14, "B": 16, "C": 10, "D": 12, "E": 12, "F": 18, "G": 14, "H": 12, "I": 18, "J": 14}
for cl, w in detail_widths.items():
    detail_ws.column_dimensions[cl].width = w
detail_ws.freeze_panes = "A2"

detail_wb.save(OUT_DETAIL)

pb_cnt2 = sum(1 for r in detail_rows if r[3] == "PB")
pe_cnt2 = sum(1 for r in detail_rows if r[3] == "PE")
print(f"\n已生成 -> {OUT_DETAIL}")
print(f"共 {len(detail_rows)} 行 | PB: {pb_cnt2}, PE: {pe_cnt2}")
